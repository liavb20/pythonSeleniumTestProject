from openpyxl import load_workbook


class Details_From_Excel:
    def __init__(self, test_number):
        # open the excel file
        self.workbook = load_workbook(filename="AOS.xlsx")
        self.sheet = self.workbook.active
        self.list_of_headers = []
        self.category, self.product_number, self.quantity, self.color = [], [], [], []
        # call the function and go to the column of its test
        self.products_details_from_excel(test_number + 2)
        # get the amount of products
        self.products_amount = len(self.category)
        self.index = -1

    # get the number of column which contain the specific test details
    def products_details_from_excel(self, col):
        # run on the column of headers of the products (category, product number, quantity, color, category, ...)
        # insert them to a list
        for val in self.sheet.iter_rows(min_row=2, max_row=13, min_col=2, max_col=2):
            self.list_of_headers.append(val[0].value)
        count = 0
        # run on the column of the specific test (column of the col) and take the product's details
        for value in self.sheet.iter_rows(min_row=2, max_row=13, min_col=col, max_col=col):
            count += 1
            # check if the cell in the excel is empty, if True stop running on the column
            if value[0].value == None:
                break
            # if the count is 1,5 or 9 the cell have the category product, add it to the category list
            if count == 1 or count == 5 or count == 9:
                x = value[0].value
                self.category.append(x)
            # if the count is 2,6 or 10 the cell have product number
            elif count == 2 or count == 6 or count == 10:
                # minus 1 from the product number because the program begin from zero and the excel from 1
                self.product_number.append(value[0].value - 1)
            # if the count 3,7 or 11 the cell have the quantity of this product, add it to list
            elif count == 3 or count == 7 or count == 11:
                self.quantity.append(value[0].value)
            # if the count is 4,8 or 12 the cell have the color for the product, add it to list
            elif count == 4 or count == 8 or count == 12:
                self.color.append(value[0].value)

    def categories(self):
        # increase the index by 1, return the category of the product
        self.index += 1
        return self.category[self.index]

    def product_numbers(self):
        # return the number of the product
        return self.product_number[self.index]

    def quantities(self):
        # return the quantity of items of the product
        return self.quantity[self.index]

    def colors(self):
        # return the color of the specific product
        return self.color[self.index]


